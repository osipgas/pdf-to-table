import pandas as pd
from pdfminer.high_level import extract_text
from openpyxl.styles import Alignment


def extract_first_page_text(pdf_path):
    text = extract_text(pdf_path, page_numbers=[0])
    return text

def get_departure_and_destination(text):
    parts_of_texts = text.split(" — ")
    departure = parts_of_texts[0]
    destination = parts_of_texts[1].split()[0]
    return departure, destination

new_cols = ["WAYPOINT", "AIRWAY", "HDG", "CRS", "ALT", "CMP", "DIR/SPD", "ISA", "TAS", "GS", "LEG", "REM", "USED", "REM", "ACT", "LEG", "REM", "ETE"]
def check_row(row, i, cols_coords):
    for j in range(len(row)):
        for tcol in new_cols:
            if type(row[j]) is str:
                col_words = row[j].strip().split()
                for col_word in col_words:
                    if col_word.strip() == tcol:
                        cols_coords[tcol] = (i, j)

def prepare_df(df):
    df = df.applymap(lambda x: x.replace('_x000D_', '') if isinstance(x, str) else x)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df = df.fillna('-')
    df.columns = df.columns.str.replace('_x000D_', '').str.strip()
    return df


def convert_to_main_table(table_path, departure, destination):
    df = pd.read_excel(table_path)
    df = prepare_df(df)
    
    table_start_row_index = 0
    table_end_row_index = len(df)

    for i, value in enumerate(df.iloc[:, 0]):
        if value == departure:
            table_start_row_index = i
        if value == destination:
            table_end_row_index = i
            
    if len(df.columns) >= 17:
        df = df.iloc[table_start_row_index:len(df), 0:18]
        df = df.iloc[:table_end_row_index, :]
        df.columns = new_cols
        cols_indices = {"WAYPOINT": 0, "ALT": 4, "HDG": 2, "CRS": 3, "DIST": 10, "TIME": 15, "EFOB": 13}
        shorten_df = df.iloc[:, [0, 4, 2, 3, 10, 15, 13]]
        shorten_df.columns = cols_indices.keys()
        return shorten_df
    else:
        return None


def convert_to_sub_table(sub_table_path):
    df = pd.read_excel(sub_table_path)
    df = prepare_df(df)

    if len(df.columns) == 10:
        new_cols = list(df.columns)
        new_cols[-1] = "LONGEST RWY LENGHT"
        new_cols[-2] = "LONGEST RWY ANGLE"
        new_cols[0] = "TYPE"

        df.columns = new_cols
        df['LONGEST RWY ANGLE'] = df['LONGEST RWY ANGLE'].astype(str).str.replace(r'[^0-9/rl]', '', regex=True)
        df = df.set_index(df.columns[0])
        return df
    else:
        return None
    




import openpyxl
from copy import copy

import pandas as pd
from tabulate import tabulate

def show_excel_table(file_path, sheet_name=0):
    """Красиво выводит Excel-таблицу в консоль"""
    # header=None — потому что у тебя первые строки — это заголовки/шапка, а не обычные колонки
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    print(f"\n=== Таблица из файла: {file_path} ===\n")
    print(tabulate(
        df, 
        headers='keys',        # нумерует столбцы как 0, 1, 2...
        tablefmt="grid",       # варианты: "grid", "pretty", "fancy_grid", "psql", "github"
        showindex=False,
        numalign="center",
        stralign="left"
    ))

import openpyxl
from copy import copy
import sys

from openpyxl.worksheet.cell_range import CellRange

def ranges_intersect(a: CellRange, b: CellRange) -> bool:
    return not (
        a.max_row < b.min_row or
        a.min_row > b.max_row or
        a.max_col < b.min_col or
        a.min_col > b.max_col
    )

def remove_intersecting_merges(ws, target_range):
    target = CellRange(target_range)
    for merge in list(ws.merged_cells.ranges):
        merge_range = CellRange(merge.coord)
        if ranges_intersect(merge_range, target):
            ws.unmerge_cells(merge.coord)


def modify_excel(input_path, N, output_path):
    if N < 1:
        print("N must be at least 1.")
        return

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active  # Assuming the active sheet is the one to modify

    # Define the repeating rows (1-based indexing)
    repeat_start = 3
    repeat_end = 5
    repeat_height = repeat_end - repeat_start + 1  # 3 rows

    current_end = repeat_end

    # The template already has one set of repeating rows, so insert (N-1) additional sets
    for _ in range(N - 1):
        # Position to insert the new rows: after the current last repeat row
        insert_row = current_end + 1

        # Insert empty rows
        ws.insert_rows(insert_row, amount=repeat_height)

        # Copy row dimensions, cell values, and styles
        for row_offset in range(repeat_height):
            src_row = repeat_start + row_offset
            dest_row = insert_row + row_offset

            # Copy row height
            try:
                height = ws.row_dimensions[src_row].height
                if height is not None:
                    ws.row_dimensions[dest_row].height = height
            except KeyError:
                pass

            # Copy cells
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=src_row, column=col)
                dest_cell = ws.cell(row=dest_row, column=col)

                dest_cell.value = src_cell.value

                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

        # Copy merged cells
        new_merges = []
        offset = insert_row - repeat_start
        for merge in list(ws.merged_cells.ranges):
            if merge.min_row >= repeat_start and merge.max_row <= repeat_end:
                new_min_row = merge.min_row + offset
                new_max_row = merge.max_row + offset
                new_merge = openpyxl.worksheet.cell_range.CellRange(
                    min_row=new_min_row,
                    max_row=new_max_row,
                    min_col=merge.min_col,
                    max_col=merge.max_col
                )
                new_merges.append(new_merge)

        for new_merge in new_merges:
            ws.merge_cells(new_merge.coord)

        # Update current_end
        current_end += repeat_height

    # merging
    for i in range(5, 17 * 3, 3):
        ws.merge_cells(f"C{i}:C{i+1}")
        ws.merge_cells(f"H{i}:H{i+1}")

    remove_intersecting_merges(ws, "C3:H4")
    ws.merge_cells("C3:H4")

    remove_intersecting_merges(ws, f"C{N*3+1}:H{N*3+2}")
    ws.merge_cells(f"C{N*3+1}:H{N*3+2}")

    return wb


def insert_values_into_template(table, ws):
    upper_col_to_letter = {"ALT": "C", "HDG": "D", "DIST": "E", "EFOB": "F"}
    lower_col_to_letter = {"CRS": "D", "TIME": "E"}

    for i in range(3, len(table) * 3 + 1, 3):
        ws[f"A{i}"].value = i / 3
        ws[f"B{i}"].value = list(table["WAYPOINT"])[int(i/3 - 1)]

    for id_table, i in enumerate(range(5, len(table) * 3, 3), start=1):
        for col, letter in upper_col_to_letter.items():
            value = list(table[col])[id_table]
            if col == "DIST":
                if len(value.split()) > 1:
                    value = value.split()[0]

            ws[f"{letter}{i}"].value = value
            if col != "ALT":
                ws[f"{letter}{i}"].alignment = Alignment(horizontal="left", vertical="center")
            else:
                ws[f"{letter}{i}"].alignment = Alignment(horizontal="center", vertical="center")

        for col, letter in lower_col_to_letter.items():
            ws[f"{letter}{i+1}"].value = list(table[col])[id_table]
            ws[f"{letter}{i+1}"].alignment = Alignment(horizontal="right", vertical="center")



import fitz

def get_names(pdf_path):
    doc = fitz.open(pdf_path)
    
    # Берем только последнюю страницу
    last_page = doc[-1] 
    
    # Получаем объекты текста (словарь с координатами)
    blocks = last_page.get_text("dict")["blocks"]
    
    # Сортируем блоки по высоте (y0), чтобы текст шел сверху вниз
    blocks.sort(key=lambda b: b["bbox"][1])
    
    texts = []
    heights = []
    for b in blocks:
        if "lines" in b:  # Проверяем, что это текстовый блок
            for line in b["lines"]:
                for span in line["spans"]:
                    y_pos = span["bbox"][1]
                    text = span["text"]

                    texts.append(text)
                    heights.append(y_pos)

    filtered_texts = []
    filtered_heights = []

    for text, height in zip(texts, heights):
        if text != "Diagram Unavailable":
            filtered_texts.append(text)
            filtered_heights.append(height)

    texts = filtered_texts
    heights = filtered_heights

    texts = texts[-4:]
    heights = heights[-4:]

    maybe_departure = texts[0].strip()
    maybe_departure_name = texts[1].strip()
    maybe_destination = texts[2].strip()
    maybe_destination_name = texts[3].strip()

    return maybe_departure, maybe_departure_name, maybe_destination, maybe_destination_name
                




def fill_template(template, info):
    for key, value in info.items():
        if value is None:
            value = "_____"
        template = template.replace(key, value)
    return template


def append_workbook_below(wb_top, wb_bottom):
    ws_top = wb_top.active
    ws_bottom = wb_bottom.active

    start_row = ws_top.max_row + 1

    for row in ws_bottom.iter_rows():
        for cell in row:
            new_cell = ws_top.cell(
                row=start_row + cell.row - 1,
                column=cell.column,
                value=cell.value
            )

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
                new_cell.protection = copy(cell.protection)

    return wb_top