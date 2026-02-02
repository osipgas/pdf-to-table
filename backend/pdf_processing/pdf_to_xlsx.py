from .things import extract_first_page_text, get_departure_and_destination, convert_to_main_table, convert_to_sub_table, modify_excel, show_excel_table, insert_values_into_template, get_names, fill_template, append_workbook_below
from pathlib import Path
from .extract_table import ExtractTextTableInfoFromPDF
import zipfile
from copy import copy
from openpyxl.styles import Alignment
import openpyxl


def extract_tables(pdf_path, extract_to):
    output_path = "pdf_processing/extracted.zip"
    ExtractTextTableInfoFromPDF(pdf_path=pdf_path, output_path=output_path)
    with zipfile.ZipFile(output_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)


def pdf_to_excel(pdf_path, tables_path, save_path, template_path):
    # processing main table
    table_path = f"{tables_path}/fileoutpart1.xlsx"

    text = extract_first_page_text(pdf_path)
    departure, destination = get_departure_and_destination(text)

    main_df = convert_to_main_table(table_path, departure, destination)


    # processing sub table
    sub_table_path = f"{tables_path}/fileoutpart8.xlsx"
    if not Path(sub_table_path).exists():
        sub_table_path = f"{tables_path}/fileoutpart7.xlsx"

    sub_df = convert_to_sub_table(sub_table_path)

    tables = {"main": main_df, "sub": sub_df}







    # inserting data
    wb = modify_excel(template_path, len(tables["main"]), "result.xlsx")
    ws = wb.active
    insert_values_into_template(tables["main"], ws)





    departure_name, destination_name = None, None

    try:
        # extracting names
        maybe_departure, maybe_departure_name, maybe_destination, maybe_destination_name = get_names(pdf_path)

        if departure == maybe_departure:
            departure_name = maybe_departure_name

        if destination == maybe_destination:
            destination_name = maybe_destination_name
    except:
        pass




    # inserting special info

    departure_info = {
        "<Name>": departure_name, 
        "<Elevation>": tables["sub"].loc["DEP", "ELEV"], 
        "<ATIS>": tables["sub"].loc["DEP", "WX"],
        "<GND>": tables["sub"].loc["DEP", "GND"],
        "<TWR>": tables["sub"].loc["DEP", "TWR/CTAF"],
        "<RWY>": tables["sub"].loc["DEP", "LONGEST RWY ANGLE"]
    }



    destination_info = {
        "<Name>": destination_name, 
        "<Elevation>": tables["sub"].loc["DEST", "ELEV"], 
        "<ATIS>": tables["sub"].loc["DEST", "WX"],
        "<GND>": tables["sub"].loc["DEST", "GND"],
        "<TWR>": tables["sub"].loc["DEST", "TWR/CTAF"],
        "<RWY>": tables["sub"].loc["DEST", "LONGEST RWY ANGLE"]
    }


    text_template = """Name: <Name>, Elevation: <Elevation>, DA: ___, Circuit pattern altitude: ___;
    ATIS: <ATIS>, GND: <GND>, TWR: <TWR>, A/A: _____ Approach: _____;
    RWY: <RWY>, Length: _____ m., Req. Dist.: ____ m., Surface: __________;
    Exp. Wind: _________, Exp. QNH: ____ hPa; Exp. TWY: ____
    RWY: ____, Wind: _______, QNH: _______, Squak: _______"""


    departure_text = fill_template(text_template, departure_info)
    destination_text = fill_template(text_template, destination_info)

    # Inserting descriptions
    ws[f"C{3}"].value = departure_text
    ws[f"C{len(main_df) * 3 + 1}"].value = destination_text

    # formating descriptions

    ws[f"C{3}"].value
    ws[f"C{3}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"C{len(main_df) * 3 + 1}"].alignment = Alignment(horizontal="left", vertical="center")


    ws.row_dimensions[3].height = 60
    ws.row_dimensions[len(main_df) * 3 + 1].height = 60


    # inserting additional text and formatting
    ws.row_dimensions[len(main_df) * 3 + 3].height = 50
    ws.merge_cells(f"A{len(main_df) * 3 + 3}:E{len(main_df) * 3 + 3}")
    ws.merge_cells(f"F{len(main_df) * 3 + 3}:H{len(main_df) * 3 + 3}")

    text_for_a = """Waypoint: Top, Track, Altitude, Radio, Engine, Estimates, Area
Diversion: Endurance, Terrain, Infrastructure, Weather, Airport
Arrival Briefing (Treats, RWY, Top Of Descent, Integration, Missed
aproach holding time, Landing config and speed, Taxiway, Apron)"""

    text_for_g = """After T/O: Flaps, Lights, Engine
Approach: QNH, Mixture, Fuel, Flaps
Landing: Mixture, Flaps, Lights
After Landing: Heat, Light, Flaps"""

    ws[f"A{len(main_df) * 3 + 3}"].value = text_for_a
    ws[f"F{len(main_df) * 3 + 3}"].value = text_for_g
    
    # saving
    wb.save(save_path)